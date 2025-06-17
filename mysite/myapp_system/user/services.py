from datetime import datetime
from openpyxl import load_workbook

from mars_framework.utils.excel import create_excel_workbook


def get_user_import_template_workbook():
    """
    生成用户导入模板的 Excel 文件。
    """

    # 定义示例数据
    data = [
        {
            "username": "xiaozhang",
            "nickname": "小张",
            "deptId": "103",
            "email": "xz@qq.com",
            "mobile": "13312345670",
            "sex": "男",
            "status": "开启",
        },
        {
            "username": "xiaoli",
            "nickname": "小李",
            "deptId": "",
            "email": "",
            "mobile": "",
            "sex": "",
            "status": "关闭",
        },
    ]

    fields_labels = {
        "username": "用户账号",
        "nickname": "用户昵称",
        "deptId": "部门编号",
        "email": "用户邮箱",
        "mobile": "手机号码",
        "sex": "用户性别",
        "status": "账号状态",
    }

    # 返回工作簿
    return create_excel_workbook(data, fields_labels)


def get_user_import_data(file) -> list:
    """
    读取Excel文件中的用户信息，并将其转换为系统所需的数据格式
    """
    # 加载Excel文件
    workbook = load_workbook(file)
    sheet = workbook.active
    # 固定表头，以确保数据字段的一致性
    headers = ["username", "nickname", "deptId", "email", "mobile", "sex", "status"]
    # 初始化数据列表
    data = []
    # 遍历每一行数据
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 将当前行数据与表头 zip 后转换为字典
        user_data = dict(zip(headers, row))
        # 转换 sex 和 status 字段
        user_data["sex"] = (
            1 if user_data["sex"] == "男" else 2 if user_data["sex"] == "女" else 0
        )
        user_data["status"] = (
            0
            if user_data["status"] == "开启"
            else 1 if user_data["status"] == "关闭" else None
        )
        data.append(user_data)

    return data


def avatar_upload_rename(instance, filename):
    """头像文件重命名，并指定存储路径"""
    ext = filename.split(".")[-1]  # 获取文件扩展名
    year = datetime.now().strftime("%Y")
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")[:-3]
    return f"avatars/{year}/{timestamp}.{ext}"  # 按年份分目录存储
