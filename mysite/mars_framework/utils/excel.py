from urllib.parse import quote
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def convert_datetime(datetime_str):
    # 移除时区部分（+08:00）
    datetime_without_tz = datetime_str.replace("+08:00", "")
    # 移除 "T" 并截取到秒部分（去掉毫秒）
    return datetime_without_tz.replace("T", " ").split(".")[0]


def process_item(item, fields, data_map={}):
    """将单个数据项转换为行数据"""
    row = []
    for field in fields:
        value = item.get(field)
        # 友好显示转换
        if field in data_map:
            value = data_map[field].get(value, None)
        # 特殊字段类型转换
        elif field == "id" or field == "job_id":
            value = str(value)
        elif field == "create_time" or field == "date_created" or field == "date_done":
            value = convert_datetime(value)
        row.append(value)
    return row


def create_excel_workbook(data, fields_labels, data_map={}, sheet_name="Sheet1"):
    """
    将数据写入Excel文件
    """
    fields = list(fields_labels.keys())
    labels = list(fields_labels.values())
    # 创建一个 Excel 工作簿
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    # 设置表头
    sheet.append(labels)
    # 添加数据
    for item in data:
        row = process_item(item, fields, data_map)
        sheet.append(row)
    # 设置列宽
    for col_num in range(1, len(labels) + 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 15

    return workbook


def generate_excel_response(workbook, file_name="export.xlsx"):
    """
    生成 Excel 文件并返回 HTTP 响应。
    """
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={quote(file_name)}"
    workbook.save(response)
    return response
